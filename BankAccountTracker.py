import pandas as pd
from openpyxl import workbook, load_workbook
import datetime
from openpyxl.utils import get_column_letter

"""
This is my first programming project that I still use 
Do not look at it.
"""


#link your xlsx file here make sure its a string (on mac right click on desired file hold down options and hit copy file path)
file = '/Users/charlesjennings/Desktop/bank_accounts/bank_account_data.xlsx'
file_2 = '/Users/charlesjennings/Desktop/myproject/Personal_Finance/management/initialization_data/Initialization_data.xlsx'
#CHANGE DESIRED SHEET NAMES HERE will interfere with data if changed in excel
sheet1 = 'overall_bank_accounts'
sheet2 = 'excel_calculations'
sheet3 = 'records'






#Checks and ensures the file is formatted correctly for program to read 
excel_file = pd.ExcelFile(file, engine = 'openpyxl')
sheet_names = excel_file.sheet_names
workbook = load_workbook(file)
if sheet_names != [sheet1, sheet2, sheet3]:
    if sheet_names[0] != sheet1:
        workbook.create_sheet(title = sheet1, index = 0)
    if sheet_names != sheet2:
        workbook.create_sheet(title = sheet2, index = 1)
    if sheet_names != sheet3:
        workbook.create_sheet(title = sheet3, index = 2)
    for i in range(len(workbook.sheetnames) - 1, -1, -1):
        if workbook.sheetnames[i] != sheet1 and workbook.sheetnames[i] != sheet2 and workbook.sheetnames[i] != sheet3:
            workbook.remove(workbook[workbook.sheetnames[i]])
    workbook.save(file)




#functions
def input_validation(userInput, lower_bound, upper_bound):
    while  (userInput < lower_bound and lower_bound != -500) or (userInput > upper_bound and upper_bound != -500):
        userInput = float(input('Error, invalid input. Please try again: '))
    return userInput    

def create_account(overall_bank_accounts):
    account_holder = input('Who is the account holder (Trin/Drew): ')
    bank_name = input('What is the name of the bank holding the account: ')
    account_type = input('What is the type of account: ')
    account_balance = float(input('What is the initial balance of the account: '))
    account_interest_rate = float(input('What is the accounts interest rate: '))
    compound_interest = input('Does interest compound (Y/N): ')
    if compound_interest == 'Y' or compound_interest == 'y':
        compound_interest_frequency = input('How often does the interest compound: ')
    else:
        compound_interest_frequency = 'N/A'
    interest_payment_frequency = input('At what time intervals is interest accrued added to the account: ')  
    routing_number = input('What is the bank routing number: ')
    bank_account_number = input('What is the account number: ')
    new_account = pd.Series([account_holder, bank_name, account_type, account_balance,
                              account_interest_rate, compound_interest, compound_interest_frequency, interest_payment_frequency, 
                              routing_number, bank_account_number])
    overall_bank_accounts = pd.concat([overall_bank_accounts, new_account.to_frame().T], ignore_index=True)
    return overall_bank_accounts

def deposits(overall_bank_accounts, overall_bank_accounts_readable, check):
    print('Input 0 at any time to return to the main menu.')
    print(overall_bank_accounts_readable)
    index_column = overall_bank_accounts.reset_index()['index']
    userInput = int(input('Please select the corresponding number to the left of the account that you moved the money into: '))
    userInput = input_validation(userInput, 0, (len(index_column) - 1 + check))
    if userInput != 0:
        num_deposit = float(input('How much did you deposit into the account: '))
        num_deposit = input_validation(num_deposit, 0, -500)
        if check == 1:
            userInput -= 1
        overall_bank_accounts.at[userInput, 3] += num_deposit
    else:
        num_deposit = 0
    return overall_bank_accounts, num_deposit, userInput

def withdrawal(overall_bank_accounts,overall_bank_accounts_readable, check):
    print('Input 0 at any time to return to the main menu.')
    print(overall_bank_accounts_readable)
    index_column = overall_bank_accounts.reset_index()['index']
    userInput = int(input('Please select the corresponding number to the left of the account that you withdrew money from: '))
    userInput = input_validation(userInput, 0, (len(index_column) - 1 + check))
    if userInput != 0:
        num_withdrawal = float(input('How much did you withdraw from the account: '))
        num_withdrawal = input_validation(num_withdrawal, 0, overall_bank_accounts.at[userInput, 3])
        if check == 1:
            userInput -= 1
        overall_bank_accounts.at[userInput, 3] -= num_withdrawal
    else:
        num_withdrawal = 0
    return overall_bank_accounts, num_withdrawal, userInput

def transfer(overall_bank_accounts, overall_bank_accounts_readable, check):
    print('Input 0 at anytime to return to the main menu.')
    print(overall_bank_accounts_readable)  
    withdrawalInput = int(input('Please enter the number to the left of the account you are transferring FROM: '))
    index_column = overall_bank_accounts.reset_index()['index']
    withdrawalInput = input_validation(withdrawalInput, 0, (len(index_column) - 1 + check))
    if withdrawalInput != 0:
        if check == 1:
            withdrawalInput -= 1
        depositInput = int(input('Please enter the number to the left of the account you are transferring TO: '))
        while depositInput == withdrawalInput:
            depositInput = int(input('Error, cannot be the same account you are transferring from. Please enter a valid account: '))
        depositInput = input_validation(depositInput, 0, (len(index_column) - 1 + check))
        if depositInput != 0:
            if check == 1:
                depositInput -= 1
            amount = float(input('Enter the amount that you are transferring: '))
            amount = input_validation(amount, 0, (overall_bank_accounts.at[withdrawalInput, 3]))
            if amount > 0:
                overall_bank_accounts.at[withdrawalInput, 3] -= amount
                overall_bank_accounts.at[depositInput, 3] += amount
    return overall_bank_accounts, withdrawalInput, depositInput, amount

def close_account(overall_bank_accounts_readable, overall_bank_accounts):
    print('Enter 0 at anytime to return back to the main menu.')
    print(overall_bank_accounts_readable)
    userInput = int(input('Please select the number to the left of the account you wish to close: '))
    if userInput != 0:
        hold = overall_bank_accounts.at[userInput, 3]
        if hold != 0:
            choice = int(input('Please select the number to the left of the account you wish to transfer the remaining balance of ' + str(hold) + ': '))
            while choice == userInput and choice != 0:
                choice = int(input('Error, you selected the account you are trying to close. Please choose a valid option: '))
            overall_bank_accounts.at[choice, 3] += hold
            overall_bank_accounts = overall_bank_accounts.drop(userInput)
            overall_bank_accounts = overall_bank_accounts.reset_index(drop = True)
    return overall_bank_accounts

def records(excel_records, userInput, num_deposit, num_withdrawal, date, overall_bank_accounts):
    if num_deposit == 0:
        withdrawal = 'Y'
        deposit = 'N'
        amount = num_withdrawal
    if num_withdrawal == 0:
        deposit = 'Y'
        withdrawal = 'N'
        amount = num_deposit
    name = overall_bank_accounts.at[userInput, 0]
    bank_name = overall_bank_accounts.at[userInput, 1]
    account_type = overall_bank_accounts.at[userInput, 2]
    new_record = pd.Series([str(date), deposit, withdrawal, amount, name, bank_name, account_type])
    excel_records = pd.concat([excel_records, new_record.to_frame().T], ignore_index = True)
    return excel_records

def calculations(overall_bank_accounts, excel_calculations,columns, date):
    total, trin_total, drew_total, j = 0, 0, 0, 0
    if overall_bank_accounts.empty == False:
        if columns == False:
            overall_bank_accounts = overall_bank_accounts.drop(0)
            j = 1
        total = overall_bank_accounts[3].sum(min_count = 1)
        print('Your overall total balance is ' + str(total))
        for i in range(len(overall_bank_accounts)):
            if overall_bank_accounts.at[(j+i), 0] == 'Trin':
                trin_total += overall_bank_accounts.at[(j+i), 3]
            if overall_bank_accounts.at[(j+i), 0] == 'Drew':
                drew_total += overall_bank_accounts.at[(j+i), 3]
    new_calculations = pd.Series([str(date), total, drew_total, trin_total])
    excel_calculations = pd.concat([excel_calculations, new_calculations.to_frame().T], ignore_index=True)
    return excel_calculations






#Declarations
columns = ['Account Holder\'s Name', 'Bank Name', 'Account Type', 'Account Balance',
            'Interest Rate', 'Compound', 'Compounding interest frequency', 'Interest Payout Frequency',
              'Routing Number', 'Account Number']
columns_readable = ['Name', 'Bank', 'Account', 'Balance']
columns_calculations = ['Date', 'Overall Total', 'Drew\'s Total', 'Trin\'s Total']
columns_records = ['Date', 'Deposiot', 'Withdrawal', 'Amount',
                    'Name', 'Bank Name', 'Account']
check = 0      #Fixes accounts readable columns bug when running program with a blank excel sheet (description: when loading
#in a blank excel file there is no index for dataframe but the readable version always has index creating an off by 1 error)
date = datetime.date.today()
#Pulls data from the linked excel file and turns it into multiple dataframes
overall_bank_accounts = pd.read_excel(file, sheet_name = sheet1, header = None, keep_default_na = False, na_filter = False)
excel_calculations = pd.read_excel(file, sheet_name = sheet2, header = None, keep_default_na = False, na_filter = False)
excel_records = pd.read_excel(file, sheet_name = sheet3, header = None, keep_default_na = False, na_filter = False)
userInput = -1
if overall_bank_accounts.empty == True:
    check += 1
#Prevents Repeating Header bug
if overall_bank_accounts.empty == False and overall_bank_accounts.iloc[0][0] == columns[0]:
    columns = False
if excel_calculations.empty == False and excel_calculations.iloc[0][0] == columns_calculations[0]:
    columns_calculations = False
if excel_records.empty == False and excel_records.iloc[0][0] == columns_records[0]:
    columns_records = False

#main menu
while (userInput != 0):
    print('MAIN MENU')
    print('1. Create account')
    print('2. Make a deposit')
    print('3. Make a withdrawal')
    print('4. Transfer money')
    print('5. close an existing account')
    print('0. exit the program')
    userInput = int(input('Please select an option: '))
    #Input validation
    while userInput >= 6 or userInput <= -1:    
        userInput = int(input('Error invalid item selected, please try again: '))
    #formats dataframe into an easy to read version for user selection specifically for visual studio code; can remove if using jupiter notebooks
    if overall_bank_accounts.empty == False and check == 0:     
        overall_bank_accounts_readable = overall_bank_accounts.drop(columns = [4, 5, 6, 7, 8, 9])
        overall_bank_accounts_readable.iloc[0] = columns_readable
    if overall_bank_accounts.empty == False and check == 1:
        overall_bank_accounts_readable = pd.DataFrame([columns_readable])
        overall_bank_accounts_readable = pd.concat([overall_bank_accounts_readable, overall_bank_accounts.drop(columns = [4, 5, 6, 7, 8, 9])])
        overall_bank_accounts_readable = overall_bank_accounts_readable.reset_index(drop = True)

    
    
    
    #Menu logic Directory
    if userInput == 1:
        overall_bank_accounts = create_account(overall_bank_accounts)
    if userInput == 2:
        if overall_bank_accounts.empty == True:      #Checks if theres an open account
            print('Sorry you do not currently have any open accounts')
        else:
            overall_bank_accounts, num_deposit, depositInput = deposits(overall_bank_accounts, overall_bank_accounts_readable, check)
            if num_deposit > 0:                  
            #if deposit function successfully executes this transfers the relevant information to the records function
                excel_records = records(excel_records, depositInput, num_deposit, 0, date, overall_bank_accounts)
            num_deposit = 0
    if userInput == 3:
        if overall_bank_accounts.empty == True:        #Checks if theres an open account
            print('Sorry you do not currently have any open accounts')
        else:
            overall_bank_accounts, num_withdrawal, withdrawalInput = withdrawal(overall_bank_accounts, overall_bank_accounts_readable, check)
            if num_withdrawal > 0:
                #if withdrawal function successfully executes this transfers the relevant information to the records function
                excel_records = records(excel_records, withdrawalInput, 0, num_withdrawal, date, overall_bank_accounts)
            num_withdrawal = 0
    if userInput == 4:
        if (len(overall_bank_accounts) < 3 and check == 0) or (len(overall_bank_accounts) < 4 and check == 1):             
            print('Error, not enough accounts documented for a transfer to occur.')
        else:
            overall_bank_accounts, withdrawalInput, depositInput, amount = transfer(overall_bank_accounts, overall_bank_accounts_readable, check)
            excel_records = records(excel_records, withdrawalInput, 0, amount, date, overall_bank_accounts)
            excel_records = records(excel_records, depositInput, amount, 0, date, overall_bank_accounts)
    if userInput == 5:
        if overall_bank_accounts.empty == True:       #Checks if theres an open account
            print('Sorry you do not currently have any open accounts')
        else:
            overall_bank_accounts = close_account(overall_bank_accounts_readable, overall_bank_accounts)
 


#Updates calculations for newly added data
if overall_bank_accounts.empty == False:
    excel_calculations = calculations(overall_bank_accounts, excel_calculations, columns, date)

#saves all data to file when program closes
with pd.ExcelWriter(file, mode = 'a', engine = "openpyxl", if_sheet_exists = 'replace',) as writer:
    if overall_bank_accounts.empty == False:
        overall_bank_accounts.to_excel(writer, sheet_name = sheet1, header = columns, index = False)
        excel_calculations.to_excel(writer, sheet_name = sheet2, header = columns_calculations, index = False)
    if excel_records.empty == False:
        excel_records.to_excel(writer, sheet_name = sheet3, header = columns_records, index = False)
    for sheet in [sheet1, sheet2, sheet3]:
        worksheet = writer.sheets[sheet]
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)  
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

with pd.ExcelWriter(file_2, mode = 'a', engine = "openpyxl", if_sheet_exists = 'replace',) as writer:
    if overall_bank_accounts.empty == False:
        overall_bank_accounts.to_excel(writer, sheet_name = sheet1, header = columns, index = False)
        excel_calculations.to_excel(writer, sheet_name = sheet2, header = columns_calculations, index = False)
    if excel_records.empty == False:
        excel_records.to_excel(writer, sheet_name = sheet3, header = columns_records, index = False)

    #Autofits Columns in the excel file
    for sheet in [sheet1, sheet2, sheet3]:
        worksheet = writer.sheets[sheet]
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)  
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
